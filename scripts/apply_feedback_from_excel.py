#!/usr/bin/env python3
import json
import os
import re
import zipfile
import subprocess
from collections import defaultdict

from openpyxl import load_workbook

ROOT = '/root/.openclaw/workspace/bphh_docgen'
CATALOG_PATH = os.path.join(ROOT, 'catalog.json')
DEMO_PATH = os.path.join(ROOT, 'project-data.demo.friendly.json')
TEMPLATES_DIR = os.path.join(ROOT, 'templates')
BUNDLE_PATH = os.path.join(ROOT, 'assets', 'templates_bundle.zip')

# default inbound file from current chat
DEFAULT_FEEDBACK_XLSX = '/root/.openclaw/media/inbound/project-data-mau_7.4---eebea859-1bd5-4f44-ad0a-9b064f09de6e.xlsx'


def norm(s: str) -> str:
    return re.sub(r'\s+', ' ', (s or '').strip().lower())


def normalize_label_for_match(s: str) -> str:
    s = norm(s)
    # bỏ phần chú thích trong ngoặc để match mềm hơn
    s = re.sub(r'\([^)]*\)', '', s)
    s = re.sub(r"\[[^\]]*\]", '', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s


def extract_new_label(note: str):
    # ưu tiên lấy text trong dấu ngoặc kép
    m = re.search(r'["“”]\s*([^"“”]+?)\s*["“”]', note)
    if m:
        return m.group(1).strip()
    # fallback: sau cụm "đổi tên"
    m = re.search(r'đổi\s*tên(?:\s*thành)?\s*[:\-]?\s*(.+)$', note, flags=re.I)
    if m:
        return m.group(1).strip(' .;:-')
    return None


def list_docx_templates():
    return sorted([f for f in os.listdir(TEMPLATES_DIR) if f.lower().endswith('.docx')])


def replace_placeholder_in_docx(path, old_key, new_key):
    old_tag = '{{' + old_key + '}}'
    new_tag = '{{' + new_key + '}}'
    changed = False

    with zipfile.ZipFile(path, 'r') as zin:
        items = {n: zin.read(n) for n in zin.namelist()}

    for name, data in list(items.items()):
        if not (name.startswith('word/') and name.endswith('.xml')):
            continue
        text = data.decode('utf-8', 'ignore')
        if old_tag in text:
            text = text.replace(old_tag, new_tag)
            items[name] = text.encode('utf-8')
            changed = True

    if changed:
        with zipfile.ZipFile(path, 'w', zipfile.ZIP_DEFLATED) as zout:
            for n, b in items.items():
                zout.writestr(n, b)
    return changed


def has_placeholder_in_docx(path, key):
    tag = '{{' + key + '}}'
    with zipfile.ZipFile(path, 'r') as zin:
        for name in zin.namelist():
            if not (name.startswith('word/') and name.endswith('.xml')):
                continue
            text = zin.read(name).decode('utf-8', 'ignore')
            if tag in text:
                return True
    return False


def rebuild_bundle():
    os.makedirs(os.path.dirname(BUNDLE_PATH), exist_ok=True)
    with zipfile.ZipFile(BUNDLE_PATH, 'w', zipfile.ZIP_DEFLATED) as z:
        z.write(CATALOG_PATH, arcname='catalog.json')
        for tpl in list_docx_templates():
            z.write(os.path.join(TEMPLATES_DIR, tpl), arcname=f'templates/{tpl}')


def refresh_labels_from_catalog_via_main_process():
    # main.cjs có logic heal label từ catalog.json gốc => chạy 1 vòng để sửa label ở runtime catalog
    cmd = ['node', '-e', "const m=require('./electron/main.cjs'); if(m&&m.__noop){}"]
    # Không gọi trực tiếp được hàm đóng kín trong main.cjs; thay vào đó chỉ ghi chú để caller tự chạy refreshRuntimeAssets khi mở app.
    return


def backup_file(path):
    b = path + '.bak'
    if not os.path.exists(b):
        with open(path, 'rb') as src, open(b, 'wb') as dst:
            dst.write(src.read())


def main(xlsx_path=DEFAULT_FEEDBACK_XLSX):
    if not os.path.exists(xlsx_path):
        raise FileNotFoundError(xlsx_path)

    with open(CATALOG_PATH, 'r', encoding='utf-8') as f:
        catalog = json.load(f)

    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb['🌐 Thông tin chung']

    # map label/sample -> key trong GLOBAL
    label_to_key = {}
    for k, v in catalog.items():
        if v.get('owner') != 'GLOBAL':
            continue
        lbl = v.get('label', '')
        smp = v.get('sample', '')
        for token in (lbl, smp):
            nk = normalize_label_for_match(token)
            if nk and nk not in label_to_key:
                label_to_key[nk] = k

    row_order_keys = []
    rename_map = {}  # key -> new label
    remove_keys = set()
    numeric_groups = defaultdict(list)  # group number -> [keys]
    unmatched = []

    for r in range(3, ws.max_row + 1):
        label = ws.cell(r, 1).value
        note = ws.cell(r, 3).value

        if label is None:
            continue
        label_s = str(label).strip()
        if not label_s:
            continue

        key = label_to_key.get(normalize_label_for_match(label_s))
        if not key:
            # fallback: tìm chứa nhau theo chuỗi đã normalize
            nlabel = normalize_label_for_match(label_s)
            for t, k2 in label_to_key.items():
                if nlabel and (nlabel in t or t in nlabel):
                    key = k2
                    break
        if not key:
            unmatched.append((r, label_s, note))
            continue

        row_order_keys.append(key)
        if note is None:
            continue

        # numeric group
        if isinstance(note, (int, float)):
            g = str(int(note))
            numeric_groups[g].append(key)
            continue

        note_s = str(note).strip()
        if not note_s:
            continue

        low = note_s.lower()
        if low == 'ok':
            continue

        if 'bỏ' in low:
            remove_keys.add(key)
            continue

        if 'đổi tên' in low:
            new_label = extract_new_label(note_s)
            if new_label:
                rename_map[key] = new_label
            continue

    # apply rename first
    for k, new_label in rename_map.items():
        if k in catalog:
            catalog[k]['label'] = new_label

    # apply numeric merges
    merged_into = {}
    tpl_files = list_docx_templates()

    key_pos = {k: i for i, k in enumerate(row_order_keys)}

    for g, keys in numeric_groups.items():
        uniq = []
        for k in keys:
            if k not in uniq:
                uniq.append(k)
        if len(uniq) < 2:
            continue

        # chọn key theo thứ tự xuất hiện feedback
        uniq.sort(key=lambda k: key_pos.get(k, 10**9))
        # chọn canon còn tồn tại trong catalog
        canon = next((k for k in uniq if k in catalog), None)
        if not canon:
            continue

        for old in uniq:
            if old == canon or old not in catalog:
                continue
            if canon not in catalog:
                break
            if catalog[old].get('owner') != catalog[canon].get('owner'):
                continue

            # replace placeholders in all templates
            for tpl in tpl_files:
                replace_placeholder_in_docx(os.path.join(TEMPLATES_DIR, tpl), old, canon)

            # merge metadata
            files = set(catalog[canon].get('files', [])) | set(catalog[old].get('files', []))
            catalog[canon]['files'] = sorted(files)
            catalog[canon]['count'] = int(catalog[canon].get('count', 0)) + int(catalog[old].get('count', 0))
            if not catalog[canon].get('sample') and catalog[old].get('sample'):
                catalog[canon]['sample'] = catalog[old]['sample']

            merged_into[old] = canon
            del catalog[old]

    # remove keys marked "bỏ" if not used in templates
    removed_real = []
    removed_skipped_used = []
    for k in sorted(remove_keys):
        if k not in catalog:
            continue
        used = False
        for tpl in tpl_files:
            if has_placeholder_in_docx(os.path.join(TEMPLATES_DIR, tpl), k):
                used = True
                break
        if used:
            removed_skipped_used.append(k)
            continue
        del catalog[k]
        removed_real.append(k)

    # update demo json according to merge/remove/rename
    with open(DEMO_PATH, 'r', encoding='utf-8') as f:
        demo = json.load(f)

    # global fields
    new_global = []
    seen_global = set()
    for item in demo.get('global_fields', []):
        key = item.get('key')
        if key in merged_into:
            key = merged_into[key]
        if key in removed_real:
            continue
        if key not in catalog:
            continue
        if key in seen_global:
            continue
        seen_global.add(key)
        new_global.append({
            'key': key,
            'label': catalog[key].get('label') or key,
            'value': item.get('value', ''),
        })

    # ensure any catalog GLOBAL key exists in demo
    exist_keys = {x['key'] for x in new_global}
    for k, v in catalog.items():
        if v.get('owner') == 'GLOBAL' and k not in exist_keys:
            new_global.append({'key': k, 'label': v.get('label') or k, 'value': v.get('sample', '')})

    demo['global_fields'] = new_global

    # jobs field keys/labels
    for job in demo.get('jobs', []):
        fields = job.get('fields', [])
        acc = {}
        order = []
        for f in fields:
            key = f.get('key')
            if key in merged_into:
                key = merged_into[key]
            if key in removed_real:
                continue
            if key not in catalog:
                continue
            val = f.get('value', '')
            if key not in acc:
                acc[key] = val
                order.append(key)
            else:
                if not acc[key] and val:
                    acc[key] = val
        job['fields'] = [
            {'key': k, 'label': catalog[k].get('label') or k, 'value': acc[k]}
            for k in order
        ]

    # write outputs
    with open(CATALOG_PATH, 'w', encoding='utf-8') as f:
        json.dump(catalog, f, ensure_ascii=False, indent=2)
        f.write('\n')

    with open(DEMO_PATH, 'w', encoding='utf-8') as f:
        json.dump(demo, f, ensure_ascii=False, indent=2)
        f.write('\n')

    rebuild_bundle()

    print('Applied feedback from Excel:', xlsx_path)
    print('Renamed:', len(rename_map))
    print('Merged keys:', len(merged_into))
    print('Removed keys:', len(removed_real))
    print('Skipped remove (still used in template):', len(removed_skipped_used))
    if unmatched:
        print('Unmatched labels:', len(unmatched))
        for r, lb, nt in unmatched[:20]:
            print(f'  row {r}: {lb} | note={nt}')


if __name__ == '__main__':
    main()
