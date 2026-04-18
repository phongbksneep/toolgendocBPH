#!/usr/bin/env python3
"""
Tạo file Excel mẫu BPHH DocGen có style đẹp để khách hàng điền dữ liệu.
Output: ../assets/project-data-mau.xlsx
"""
import json, os, sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                               GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_DATE_DDMMYY

SCRIPT_DIR = Path(__file__).parent
ROOT_DIR   = SCRIPT_DIR.parent
CATALOG    = ROOT_DIR / "catalog.json"
DEMO_JSON  = ROOT_DIR / "project-data.demo.friendly.json"
OUT_FILE   = ROOT_DIR / "assets" / "project-data-mau.xlsx"

with open(CATALOG, encoding="utf-8") as f:
    catalog = json.load(f)
with open(DEMO_JSON, encoding="utf-8") as f:
    demo = json.load(f)

# ── Color palette ──────────────────────────────────────────────────────────────
C_NAVY      = "FF1F3864"   # dark navy blue  – sheet titles / main headers
C_BLUE      = "FF2E75B6"   # medium blue     – sub-headers
C_LBLUE     = "FFD6E4F0"   # light blue      – alternate data rows
C_GOLD      = "FFBF8F00"   # gold            – highlight key column
C_LGOLD     = "FFFFF2CC"   # light gold      – alternate key rows
C_GREEN     = "FF375623"   # dark green      – section titles (NTCV)
C_LGREEN    = "FFE2EFD9"   # light green
C_BROWN     = "FF843C0C"   # brown           – VAT_LIEU section
C_LBROWN    = "FFFCE4D6"   # light brown
C_PURPLE    = "FF4B0082"   # purple          – YC_NTCV section
C_LPURPLE   = "FFE8DCFA"   # light purple
C_WHITE     = "FFFFFFFF"
C_KEY_GRAY  = "FFD9D9D9"   # key row background
C_BORDER    = "FFB0C4DE"
C_GUIDE_BG  = "FF172B4D"   # guide sheet bg

# ── Helpers ───────────────────────────────────────────────────────────────────
def fill(argb):
    return PatternFill("solid", fgColor=argb)

def font(bold=False, italic=False, size=10, color=C_WHITE, name="Calibri"):
    return Font(bold=bold, italic=italic, size=size, color=color, name=name)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def thin_border():
    s = Side(style="thin", color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)

def set_cell(ws, row, col, value, fnt=None, fil=None, aln=None, brd=None):
    c = ws.cell(row=row, column=col, value=value)
    if fnt: c.font = fnt
    if fil: c.fill = fil
    if aln: c.alignment = aln
    if brd: c.border = brd
    return c

def get_demo_global():
    """Return dict of key→value from demo global_fields."""
    result = {}
    for entry in demo.get("global_fields", []):
        result[entry["key"]] = entry.get("value", "")
    return result

def get_demo_jobs(owner):
    """Return list of field dicts for jobs belonging to owner."""
    result = []
    for job in demo.get("jobs", []):
        fields = job.get("fields", {})
        if isinstance(fields, list):
            fd = {x["key"]: x.get("value", "") for x in fields}
        else:
            fd = fields
        if fd and any(k.startswith(owner + "__") for k in fd):
            result.append(fd)
    return result

# ── Workbook ──────────────────────────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)  # remove default sheet

# ════════════════════════════════════════════════════════════════════════════════
# Sheet 1 – Hướng dẫn
# ════════════════════════════════════════════════════════════════════════════════
ws_guide = wb.create_sheet("📖 Hướng dẫn")
ws_guide.sheet_properties.tabColor = "1F3864"
ws_guide.column_dimensions["A"].width = 4
ws_guide.column_dimensions["B"].width = 38
ws_guide.column_dimensions["C"].width = 58
ws_guide.column_dimensions["D"].width = 20

# Title banner
ws_guide.merge_cells("A1:D1")
set_cell(ws_guide, 1, 1, "BPHH DocGen – File Nhập Liệu Mẫu",
         fnt=font(bold=True, size=16, color=C_WHITE),
         fil=fill(C_NAVY),
         aln=align("center"))
ws_guide.row_dimensions[1].height = 36

ws_guide.merge_cells("A2:D2")
set_cell(ws_guide, 2, 1,
         "Điền dữ liệu vào các sheet bên dưới rồi import vào BPHH DocGen",
         fnt=font(size=11, color="FFADD8E6"),
         fil=fill(C_GUIDE_BG),
         aln=align("center"))
ws_guide.row_dimensions[2].height = 22

GUIDE_ROWS = [
    ("", ""),
    ("Sheet",          "Nội dung"),
    ("🌐 Thông tin chung",  "Các trường dùng chung cho tất cả biên bản – điền vào cột Giá trị"),
    ("📄 BB Nghiệm thu CV", "Mỗi hàng = 1 biên bản nghiệm thu công việc"),
    ("🧱 Nghiệm thu VL",    "Mỗi hàng = 1 biên bản nghiệm thu vật liệu"),
    ("✅ Yêu cầu NTCV",     "Mỗi hàng = 1 yêu cầu nghiệm thu công việc"),
    ("", ""),
    ("Lưu ý",          "Không sửa tên sheet, không xóa hàng tiêu đề màu xanh"),
    ("Lưu ý",          "Hàng màu xám nhạt chứa tên biến (key) – KHÔNG điền vào hàng này"),
    ("Lưu ý",          "Lưu file dưới định dạng .xlsx trước khi import"),
    ("Lưu ý",          "Dùng menu 📥 Import Excel trong app để nạp dữ liệu"),
]

for r, (col_b, col_c) in enumerate(GUIDE_ROWS, start=4):
    is_header = col_b == "Sheet"
    is_note   = col_b == "Lưu ý"
    bg = fill(C_BLUE) if is_header else (fill("FFFFFFFF") if not col_b else fill("FFF5F5F5"))
    fc = font(bold=is_header, size=10, color="FF333333" if not is_header else C_WHITE)
    set_cell(ws_guide, r, 2, col_b, fnt=fc, fil=bg,
             aln=Alignment(horizontal="left", vertical="center"))
    if is_note:
        fc2 = font(size=10, color="FF8B0000", italic=True)
        bg2 = fill("FFFFF8F0")
    elif is_header:
        fc2 = font(bold=True, size=10, color=C_WHITE)
        bg2 = fill(C_BLUE)
    else:
        fc2 = font(size=10, color="FF333333")
        bg2 = bg
    set_cell(ws_guide, r, 3, col_c, fnt=fc2, fil=bg2,
             aln=Alignment(horizontal="left", vertical="center", wrap_text=True))
    ws_guide.row_dimensions[r].height = 18

# ════════════════════════════════════════════════════════════════════════════════
# Sheet 2 – Thông tin chung (GLOBAL)
# ════════════════════════════════════════════════════════════════════════════════
ws_g = wb.create_sheet("🌐 Thông tin chung")
ws_g.sheet_properties.tabColor = "2E75B6"
ws_g.column_dimensions["A"].width = 42
ws_g.column_dimensions["B"].width = 40
ws_g.column_dimensions["C"].width = 7

# Title
ws_g.merge_cells("A1:B1")
set_cell(ws_g, 1, 1, "THÔNG TIN CHUNG",
         fnt=font(bold=True, size=14, color=C_WHITE),
         fil=fill(C_NAVY), aln=align("center"))
ws_g.row_dimensions[1].height = 30

# Column headers row 2
for col, txt in [(1, "Tên trường"), (2, "Giá trị")]:
    set_cell(ws_g, 2, col, txt,
             fnt=font(bold=True, size=10, color=C_WHITE),
             fil=fill(C_BLUE),
             aln=align("center"),
             brd=thin_border())
ws_g.row_dimensions[2].height = 22

global_demo = get_demo_global()
global_fields = [(k, v) for k, v in catalog.items() if v["owner"] == "GLOBAL"]

for i, (key, meta) in enumerate(global_fields):
    row = i + 3
    alt = fill(C_LBLUE) if i % 2 == 0 else fill(C_WHITE)
    demo_val = global_demo.get(key, "")

    set_cell(ws_g, row, 1, meta["label"],
             fnt=font(size=10, color="FF1F2D3D"),
             fil=alt, aln=Alignment(horizontal="left", vertical="center", wrap_text=True),
             brd=thin_border())
    vc = ws_g.cell(row=row, column=2, value=demo_val)
    vc.fill = PatternFill("solid", fgColor="FFFFFBE8") if demo_val else fill(C_WHITE)
    vc.font = Font(size=10, color="FF1A1A1A", name="Calibri")
    vc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    vc.border = thin_border()
    ws_g.row_dimensions[row].height = 18

# Freeze top 2 rows
ws_g.freeze_panes = "A3"

# ════════════════════════════════════════════════════════════════════════════════
# Helper: build a list-type sheet (NTCV / VAT_LIEU / YC_NTCV)
# ════════════════════════════════════════════════════════════════════════════════
def build_list_sheet(sheet_name, owner, tab_color,
                     header_color, alt_color, template_name):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_properties.tabColor = tab_color.lstrip("FF")

    fields = [(k, v) for k, v in catalog.items() if v["owner"] == owner]
    ncols = len(fields)

    # ── Row 1: banner ──────────────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    set_cell(ws, 1, 1,
             sheet_name.replace("📄 ", "").replace("🧱 ", "").replace("✅ ", ""),
             fnt=font(bold=True, size=13, color=C_WHITE),
             fil=fill(C_NAVY), aln=align("center"))
    ws.row_dimensions[1].height = 28

    # ── Row 2: field labels ────────────────────────────────────────────────────
    for col, (key, meta) in enumerate(fields, start=1):
        set_cell(ws, 2, col, meta["label"],
                 fnt=font(bold=True, size=9, color=C_WHITE),
                 fil=fill(header_color),
                 aln=Alignment(horizontal="center", vertical="center",
                               wrap_text=True),
                 brd=thin_border())
    ws.row_dimensions[2].height = 36

    # ── Row 3: keys (light gray – for import reference) ────────────────────────
    for col, (key, meta) in enumerate(fields, start=1):
        set_cell(ws, 3, col, key,
                 fnt=Font(size=7, italic=True, color="FF888888", name="Calibri"),
                 fil=fill("FFE8E8E8"),
                 aln=Alignment(horizontal="center", vertical="center"))
        ws.row_dimensions[3].height = 12

    # ── Rows 4+: demo data ─────────────────────────────────────────────────────
    demo_rows = get_demo_jobs(owner)
    if not demo_rows:
        # At least 3 blank rows
        demo_rows = [{} for _ in range(3)]

    for r_idx, job_fields in enumerate(demo_rows[:5]):   # max 5 sample rows
        row = r_idx + 4
        is_alt = r_idx % 2 == 0
        bg = fill(alt_color) if is_alt else fill(C_WHITE)
        for col, (key, meta) in enumerate(fields, start=1):
            val = job_fields.get(key, "")
            c = ws.cell(row=row, column=col, value=val)
            if val:
                c.fill = PatternFill("solid", fgColor="FFFFFBE8")
            elif is_alt:
                c.fill = PatternFill("solid", fgColor=alt_color)
            else:
                c.fill = PatternFill("solid", fgColor="FFFFFFFF")
            c.font = Font(size=9, color="FF1A1A1A", name="Calibri")
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            c.border = thin_border()
        ws.row_dimensions[row].height = 20

    # Empty rows for user input
    for extra in range(5):
        row = len(demo_rows) + 4 + extra
        for col in range(1, ncols + 1):
            c = ws.cell(row=row, column=col, value="")
            c.fill = PatternFill("solid", fgColor="FFFFFFFF")
            c.border = thin_border()
        ws.row_dimensions[row].height = 20

    # ── Column widths ──────────────────────────────────────────────────────────
    for col, (key, meta) in enumerate(fields, start=1):
        col_letter = get_column_letter(col)
        # Estimate width from label length
        lbl_len = len(meta["label"])
        ws.column_dimensions[col_letter].width = max(14, min(30, lbl_len * 1.3))

    ws.freeze_panes = "A4"
    return ws

build_list_sheet(
    "📄 BB Nghiệm thu CV", "LIST_NTCV",
    "FF375623", "FF375623", C_LGREEN,
    "BB Nghiệm thu công việc.docx"
)
build_list_sheet(
    "🧱 Nghiệm thu VL", "LIST_VAT_LIEU",
    "FF843C0C", "FF843C0C", C_LBROWN,
    "Nghiệm thu vật liệu.docx"
)
build_list_sheet(
    "✅ Yêu cầu NTCV", "LIST_YC_NTCV",
    "FF4B0082", "FF4B0082", C_LPURPLE,
    "Yêu cầu nghiệm thu công việc.docx"
)

# ── Save ─────────────────────────────────────────────────────────────────────
OUT_FILE.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT_FILE)
print(f"✓ Tạo Excel mẫu: {OUT_FILE}")
